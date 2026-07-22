"""Parser da tabela de preços oficial da TDP (xlsx).

Layout: 1 aba, cabeçalho com COD | Origem/Descrição | Produtor | Região |
Classificação | Safra | Cód.Barras | Embal | Volume | Valor Final. O PAÍS não é
coluna — vem de linhas-seção (BRASIL, ARGENTINA, FRANCA...). A descrição é
codificada mas traz uva/cor/tipo (ex.: "...SAUVIGNON BLANC BCO", "...MALBEC TTO",
"ESPUMANTE BRUT BCO"), que o enrichment (LLM) normaliza depois.
"""

import re
import logging

logger = logging.getLogger(__name__)


def _num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.replace("R$", "").replace(".", "").replace(",", ".").strip()
        try:
            return float(s)
        except ValueError:
            return None
    return None


def parse_catalog(xlsx_path: str) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Localiza a linha de cabeçalho (a que tem "COD" na col A)
    header_idx = 0
    for i, r in enumerate(rows):
        if r and str(r[0]).strip().upper() == "COD":
            header_idx = i
            break

    wines: list[dict] = []
    current_country = None

    for r in rows[header_idx + 1:]:
        if not r:
            continue
        col0 = str(r[0]).strip() if r[0] is not None else ""
        preco = _num(r[9]) if len(r) > 9 else None

        # Linha-seção de país: col A com texto (não começa com dígito) e sem preço
        if col0 and preco is None and not re.match(r"^\d", col0):
            # ignora notas tipo "** Validade..."
            if not col0.startswith("*"):
                current_country = col0
            continue

        # Linha de dado: precisa de COD + preço
        if not col0 or preco is None:
            continue

        wines.append({
            "sku": col0,
            "descricao_raw": str(r[1] or "").strip(),
            "produtor": str(r[2] or "").strip(),
            "regiao": str(r[3] or "").strip(),
            "classificacao": str(r[4] or "").strip(),
            "safra": str(r[5] or "").strip(),
            "codigo_barras": str(r[6] or "").strip(),
            "embalagem": str(r[7] or "").strip(),
            "volume": str(r[8] or "").strip(),
            "preco": preco,
            "pais": current_country or "",
        })

    logger.info(f"Catálogo TDP parseado: {len(wines)} vinhos")
    return wines


if __name__ == "__main__":
    import sys, glob
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    path = sys.argv[1] if len(sys.argv) > 1 else glob.glob("ONN TRADE*.xlsx")[0]
    ws = parse_catalog(path)
    print(f"{len(ws)} vinhos")
    from collections import Counter
    print("por país:", dict(Counter(w["pais"] for w in ws)))
    for w in ws[:3]:
        print(w)
